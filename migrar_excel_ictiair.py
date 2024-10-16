import os, sys
from django.core.wsgi import get_wsgi_application

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pryictiair.settings')

application = get_wsgi_application()
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from landing.models import PersonNotificacion
from pryictiair.settings import BASE_DIR
from django.db import transaction
import openpyxl
import json

wrkbk = openpyxl.load_workbook(os.path.join(BASE_DIR, "leads_conference.xlsx"))
h = wrkbk.worksheets[0]

if __name__ == '__main__':
    try:
        with transaction.atomic():
            list_no_creados = []
            for i in range(3, h.max_row + 1):
                identificacion = h.cell(row=i, column=4).value
                last_name = h.cell(row=i, column=5).value
                first_name = h.cell(row=i, column=6).value
                middle_name = h.cell(row=i, column=7).value
                name_prefix = h.cell(row=i, column=8).value
                name_suffix = h.cell(row=i, column=9).value
                prefered_name = h.cell(row=i, column=10).value
                email = h.cell(row=i, column=11).value
                country = h.cell(row=i, column=26).value
                try:
                    person = PersonNotificacion(
                        identification=identificacion,
                        last_name=last_name,
                        first_name=first_name,
                        middle_name=middle_name,
                        name_prefix=name_prefix,
                        name_suffix=name_suffix,
                        email=email,
                        country=country
                    )
                    person.save()
                    print(f"{i}.- Se agrego el registro: ",
                          [identificacion, last_name, first_name, middle_name, name_prefix, name_suffix, email,
                           country])
                except Exception as e:
                    list_no_creados.append(identificacion)
            print(f"Registros no creados({len(list_no_creados)})", list_no_creados)
    except Exception as ex:
        print("Error: " + str(ex))
